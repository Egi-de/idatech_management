from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.db.models import Sum, Q
from django.http import JsonResponse, HttpResponse
from django.middleware.csrf import get_token
from .models import Student, Employee, Expense, Transaction, RecentActivity, TrashBinEntry
from django.contrib.auth.decorators import login_required
from user_auth.models import Profile
import json
from django.utils import timezone
from django.core.serializers import serialize
from django.contrib.contenttypes.models import ContentType
from django.contrib.contenttypes.fields import GenericForeignKey
from django.template.loader import render_to_string
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
from django.core.mail import EmailMessage

@login_required
def add_employee_form(request):
    if request.method == 'GET':
        return render(request, 'admin_panel/add_employee.html')

@login_required
def dashboard(request):
    # Fetch counts and summaries for dashboard
    total_students = Student.objects.count()
    total_employees = Employee.objects.count()
    iot_students = Student.objects.filter(program='iot').count()
    sod_students = Student.objects.filter(program='sod').count()

    # Financial summaries
    total_salaries = Employee.objects.aggregate(Sum('salary'))['salary__sum'] or 0
    transport_expenses = Expense.objects.filter(type=Expense.TRANSPORT).aggregate(Sum('amount'))['amount__sum'] or 0
    other_expenses = Expense.objects.filter(type=Expense.OTHER).aggregate(Sum('amount'))['amount__sum'] or 0
    total_expenses = total_salaries + transport_expenses + other_expenses

    # Recent transactions (last 5)
    recent_transactions = Transaction.objects.order_by('-date')[:5]

    # Filtering and sorting for recent activities
    recent_activities_qs = RecentActivity.objects.all()

    # Filtering by user
    user_filter = request.GET.get('user')
    if user_filter:
        recent_activities_qs = recent_activities_qs.filter(user__username__icontains=user_filter)

    # Filtering by action text search
    action_search = request.GET.get('action_search')
    if action_search:
        recent_activities_qs = recent_activities_qs.filter(action__icontains=action_search)

    # Filtering by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        recent_activities_qs = recent_activities_qs.filter(timestamp__date__range=[start_date, end_date])

    # Sorting
    sort_by = request.GET.get('sort_by', 'timestamp_desc')
    if sort_by == 'timestamp_asc':
        recent_activities_qs = recent_activities_qs.order_by('timestamp')
    elif sort_by == 'action_asc':
        recent_activities_qs = recent_activities_qs.order_by('action')
    elif sort_by == 'action_desc':
        recent_activities_qs = recent_activities_qs.order_by('-action')
    else:
        # Default to timestamp descending
        recent_activities_qs = recent_activities_qs.order_by('-timestamp')

    # Limit to 20 for performance
    recent_activities = recent_activities_qs[:20]

    # Populate user filter dropdown with distinct usernames
    user_list = RecentActivity.objects.values_list('user__username', flat=True).distinct().order_by('user__username')

    # Fetch all students, employees, expenses for rendering
    sort = request.GET.get('sort', 'name_asc')
    if sort == 'name_asc':
        students = Student.objects.all().order_by('name')
    elif sort == 'name_desc':
        students = Student.objects.all().order_by('-name')
    elif sort == 'date_asc':
        students = Student.objects.all().order_by('enrollment_date')
    elif sort == 'date_desc':
        students = Student.objects.all().order_by('-enrollment_date')
    else:
        students = Student.objects.all().order_by('name')

    employees = Employee.objects.all()
    expenses = Expense.objects.all()

    # Get or create profile for logged-in user
    profile, created = Profile.objects.get_or_create(user=request.user)

    context = {
        'total_students': total_students,
        'total_employees': total_employees,
        'iot_students': iot_students,
        'sod_students': sod_students,
        'total_salaries': total_salaries,
        'transport_expenses': transport_expenses,
        'other_expenses': other_expenses,
        'total_expenses': total_expenses,
        'recent_transactions': recent_transactions,
        'recent_activities': recent_activities,
        'students': students,
        'employees': employees,
        'expenses': expenses,
        'profile': profile,
        'user_list': user_list,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Return JSON for AJAX GET requests (for filtering/sorting)
        if request.method == 'GET':
            activities = []
            for activity in recent_activities:
                activities.append({
                    'id': activity.id,
                    'action': activity.action,
                    'icon_class': activity.icon_class,
                    'timestamp': activity.timestamp.isoformat(),
                    'user': activity.user.username,
                })
            return JsonResponse({'activities': activities})
        # Return only the recent activities partial HTML for AJAX POST requests (legacy, if any)
        return render(request, 'admin_panel/recent_activities_partial.html', context)

    return render(request, 'admin_panel/index.html', context)

@login_required
def student_report_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    return render(request, 'admin_panel/student_report_detail.html', {'student': student})

@login_required
def employment_report_detail(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    return render(request, 'admin_panel/employment_report_detail.html', {'employee': employee})

@login_required
def export_student_report_pdf(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 50
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, y, f"Student Report: {student.name}")
    y -= 30

    p.setFont("Helvetica", 12)
    p.drawString(50, y, f"Student ID: {student.student_id}")
    y -= 20
    p.drawString(50, y, f"Email: {student.email}")
    y -= 20
    p.drawString(50, y, f"Phone: {student.phone}")
    y -= 20
    p.drawString(50, y, f"Address: {student.address}")
    y -= 30

    # Add more fields as needed...

    p.showPage()
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{student.name}_report.pdf"'
    return response

@login_required
def export_student_report_excel(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Report"

    ws.append(["Field", "Value"])
    ws.append(["Name", student.name])
    ws.append(["Student ID", student.student_id])
    ws.append(["Email", student.email])
    ws.append(["Phone", student.phone])
    ws.append(["Address", student.address])
    ws.append(["Enrollment Date", student.enrollment_date.strftime("%Y-%m-%d")])
    ws.append(["Type", student.get_type_display()])
    ws.append(["Category", student.category])
    ws.append(["Program", student.get_program_display()])
    ws.append(["Level", student.level])
    ws.append(["Total Sessions", student.total_sessions])
    ws.append(["Attended Sessions", student.attended_sessions])
    ws.append(["Absences", student.absences])
    ws.append(["Participation Score", float(student.participation_score)])
    ws.append(["Average Scores", float(student.avg_scores)])
    ws.append(["Project Completion Rate", float(student.project_completion_rate)])
    ws.append(["Certifications", student.certifications])
    ws.append(["Hackathons Attended", student.hackathons_attended])
    ws.append(["Awards", student.awards])
    ws.append(["Contributions", student.contributions])
    ws.append(["Mentor Comments", student.mentor_comments])
    ws.append(["Peer Reviews", student.peer_reviews])
    ws.append(["Strengths", student.strengths])
    ws.append(["Areas for Improvement", student.areas_for_improvement])
    ws.append(["Current Status", student.get_current_status_display()])
    ws.append(["Next Steps", student.next_steps])
    ws.append(["Graduation Eligibility", "Yes" if student.graduation_eligibility else "No"])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{student.name}_report.xlsx"'
    wb.save(response)
    return response

@login_required
def email_student_report(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        recipient_email = request.POST.get('email')
        if recipient_email:
            subject = f"Student Report: {student.name}"
            message = render_to_string('admin_panel/email_student_report.html', {'student': student})
            email = EmailMessage(subject, message, to=[recipient_email])
            email.content_subtype = "html"
            email.send()
            return JsonResponse({'success': True, 'message': 'Email sent successfully'})
        return JsonResponse({'success': False, 'message': 'Recipient email is required'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def delete_recent_activity(request, activity_id):
    if request.method == 'POST':
        activity = get_object_or_404(RecentActivity, id=activity_id)
        # Save to TrashBinEntry before deleting
        TrashBinEntry.objects.create(
            user=request.user,
            item_type='RecentActivity',
            item_id=activity.id,
            item_data={
                'action': activity.action,
                'icon_class': activity.icon_class,
                'timestamp': activity.timestamp.isoformat(),
            },
            deleted_at=timezone.now()
        )
        activity.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        return redirect('admin_panel:dashboard')
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

import uuid

@require_POST
@login_required
def add_student(request):
    def generate_unique_student_id():
        return str(uuid.uuid4())[:8]  # Short unique ID

    name = request.POST.get('student-name')
    type_ = request.POST.get('student-type')
    program = request.POST.get('student-program')
    level = request.POST.get('student-level')
    address = request.POST.get('student-address')
    if name and type_ and program and level and address:
        student_id = generate_unique_student_id()
        student = Student.objects.create(
            name=name,
            student_id=student_id,
            type=type_,
            program=program,
            level=level,
            address=address,
        )
        # Create recent activity
        RecentActivity.objects.create(
            user=request.user,
            action=f"New trainee {student.name} added to {student.get_program_display()} program",
            icon_class="fas fa-plus-circle text-green-500"
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            # AJAX request
            total_students = Student.objects.count()
            iot_students = Student.objects.filter(program='iot').count()
            sod_students = Student.objects.filter(program='sod').count()
            return JsonResponse({
                'success': True,
            'student': {
                'id': student.id,
                'name': student.name,
                'type': student.get_type_display(),
                'category': student.category,
                'address': student.address,
                'program': student.get_program_display(),
                'level': student.level,
            },
                'counts': {
                    'total_students': total_students,
                    'iot_students': iot_students,
                    'sod_students': sod_students,
                }
            })
    return redirect('admin_panel:dashboard')

@require_POST
@login_required
def add_employee(request):
    name = request.POST.get('employee-name')
    position = request.POST.get('employee-position')
    department = request.POST.get('employee-department')
    salary = request.POST.get('employee-salary')
    if name and position and department and salary:
        employee = Employee.objects.create(
            name=name,
            position=position,
            department=department,
            salary=salary,
        )
        # Create recent activity
        RecentActivity.objects.create(
            user=request.user,
            action=f"New employee {employee.name} added to {employee.department} department",
            icon_class="fas fa-user-plus text-green-500"
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            # AJAX request
            total_employees = Employee.objects.count()
            total_salaries = Employee.objects.aggregate(Sum('salary'))['salary__sum'] or 0
            transport_expenses = Expense.objects.filter(type=Expense.TRANSPORT).aggregate(Sum('amount'))['amount__sum'] or 0
            other_expenses = Expense.objects.filter(type=Expense.OTHER).aggregate(Sum('amount'))['amount__sum'] or 0
            total_expenses = total_salaries + transport_expenses + other_expenses
            return JsonResponse({
                'success': True,
                'employee': {
                    'id': employee.id,
                    'name': employee.name,
                    'position': employee.position,
                    'department': employee.department,
                    'salary': str(employee.salary),
                },
                'financial': {
                    'total_salaries': total_salaries,
                    'total_expenses': total_expenses,
                },
                'counts': {
                    'total_employees': total_employees,
                }
            })
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Missing required fields'})
    return redirect('admin_panel:dashboard')

@login_required
def update_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        student.name = request.POST.get('student-name')
        student.type = request.POST.get('student-type')
        student.program = request.POST.get('student-program')
        student.level = request.POST.get('student-level')
        student.address = request.POST.get('student-address')
        student.save()
        # Create recent activity
        RecentActivity.objects.create(
            user=request.user,
            action=f"Updated trainee {student.name} in {student.get_program_display()} program",
            icon_class="fas fa-edit text-blue-500"
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'student': {
                    'id': student.id,
                    'name': student.name,
                    'type': student.get_type_display(),
                    'category': student.category,
                    'address': student.address,
                    'program': student.get_program_display(),
                    'level': student.level,
                }
            })
        return redirect('admin_panel:dashboard')
    return redirect('admin_panel:dashboard')

@login_required
def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        # Save to TrashBinEntry before deleting
        TrashBinEntry.objects.create(
            user=request.user,
            item_type='Student',
            item_id=student.id,
            item_data=json.loads(serialize('json', [student]))[0]['fields'],
            deleted_at=timezone.now()
        )
        # Create recent activity before deleting
        RecentActivity.objects.create(
            user=request.user,
            action=f"Deleted trainee {student.name} from {student.get_program_display()} program",
            icon_class="fas fa-trash text-red-500"
        )
        student.delete()
        return redirect('admin_panel:dashboard')
    context = {'student': student}
    return render(request, 'admin_panel/delete_student.html', context)

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@require_POST
@login_required
def bulk_delete_students(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            student_ids = data.get('student_ids', [])
            if not student_ids:
                return JsonResponse({'success': False, 'error': 'No student IDs provided'})

            students = Student.objects.filter(id__in=student_ids)
            deleted_count = 0
            for student in students:
                # Save to TrashBinEntry before deleting
                TrashBinEntry.objects.create(
                    user=request.user,
                    item_type='Student',
                    item_id=student.id,
                    item_data=json.loads(serialize('json', [student]))[0]['fields'],
                    deleted_at=timezone.now()
                )
                deleted_count += 1

            # Bulk delete
            students.delete()

            # Create recent activity
            RecentActivity.objects.create(
                user=request.user,
                action=f"Bulk deleted {deleted_count} students",
                icon_class="fas fa-trash text-red-500"
            )

            return JsonResponse({'success': True, 'deleted_count': deleted_count})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required
def update_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        employee.name = request.POST.get('employee-name')
        employee.position = request.POST.get('employee-position')
        employee.department = request.POST.get('employee-department')
        employee.salary = request.POST.get('employee-salary')
        employee.save()
        # Create recent activity
        RecentActivity.objects.create(
            user=request.user,
            action=f"Updated employee {employee.name} in {employee.department} department",
            icon_class="fas fa-edit text-blue-500"
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'employee': {
                    'id': employee.id,
                    'name': employee.name,
                    'position': employee.position,
                    'department': employee.department,
                    'salary': str(employee.salary),
                }
            })
        return redirect('admin_panel:dashboard')
    return redirect('admin_panel:dashboard')

@login_required
def delete_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        # Save to TrashBinEntry before deleting
        TrashBinEntry.objects.create(
            user=request.user,
            item_type='Employee',
            item_id=employee.id,
            item_data=json.loads(serialize('json', [employee]))[0]['fields'],
            deleted_at=timezone.now()
        )
        # Create recent activity before deleting
        RecentActivity.objects.create(
            user=request.user,
            action=f"Deleted employee {employee.name} from {employee.department} department",
            icon_class="fas fa-trash text-red-500"
        )
        employee.delete()
        return redirect('admin_panel:dashboard')
    context = {'employee': employee}
    return render(request, 'admin_panel/delete_employee.html', context)

@require_POST
@login_required
def add_expense(request):
    type_ = request.POST.get('expense-type')
    description = request.POST.get('expense-description')
    amount = request.POST.get('expense-amount')
    if type_ and description and amount:
        expense = Expense.objects.create(
            type=type_,
            description=description,
            amount=amount,
        )
        # Also create a transaction for the expense
        Transaction.objects.create(
            type=type_,
            description=description,
            amount=amount,
        )
        # Create recent activity
        RecentActivity.objects.create(
            user=request.user,
            action=f"New expense added: {description} - ${amount}",
            icon_class="fas fa-dollar-sign text-green-500"
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            # AJAX request
            total_salaries = Employee.objects.aggregate(Sum('salary'))['salary__sum'] or 0
            transport_expenses = Expense.objects.filter(type=Expense.TRANSPORT).aggregate(Sum('amount'))['amount__sum'] or 0
            other_expenses = Expense.objects.filter(type=Expense.OTHER).aggregate(Sum('amount'))['amount__sum'] or 0
            total_expenses = total_salaries + transport_expenses + other_expenses
            return JsonResponse({
                'success': True,
                'total_salaries': total_salaries,
                'transport_expenses': transport_expenses,
                'other_expenses': other_expenses,
                'total_expenses': total_expenses,
            })
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Missing required fields'})
    return redirect('admin_panel:dashboard')

def recent_transactions(request):
    # Get query params for filtering, sorting, searching
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort_by', 'date_desc')
    filter_type = request.GET.get('filter_type', '')

    transactions_qs = Transaction.objects.all()

    # Filter by type if provided
    if filter_type:
        transactions_qs = transactions_qs.filter(type=filter_type)

    # Search in description and type display
    if search_query:
        transactions_qs = transactions_qs.filter(
            Q(description__icontains=search_query) |
            Q(type__icontains=search_query)
        )

    # Sorting
    if sort_by == 'date_asc':
        transactions_qs = transactions_qs.order_by('date')
    elif sort_by == 'date_desc':
        transactions_qs = transactions_qs.order_by('-date')
    elif sort_by == 'amount_asc':
        transactions_qs = transactions_qs.order_by('amount')
    elif sort_by == 'amount_desc':
        transactions_qs = transactions_qs.order_by('-amount')
    else:
        transactions_qs = transactions_qs.order_by('-date')

    # Limit to 50 for performance
    transactions = transactions_qs[:50]

    data = []
    for tx in transactions:
        data.append({
            'id': tx.id,
            'date': tx.date.strftime('%Y-%m-%d'),
            'type': tx.get_type_display(),
            'description': tx.description,
            'amount': str(tx.amount),
        })
    return JsonResponse({'transactions': data})

@login_required
@require_POST
def ajax_delete_transaction(request, transaction_id):
    try:
        transaction = Transaction.objects.get(id=transaction_id)
        # Save to TrashBinEntry before deleting
        TrashBinEntry.objects.create(
            user=request.user,
            item_type='Transaction',
            item_id=transaction.id,
            item_data=json.loads(serialize('json', [transaction]))[0]['fields'],
            deleted_at=timezone.now()
        )
        transaction.delete()
        return JsonResponse({'success': True})
    except Transaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transaction not found'})

@login_required
def update_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    if request.method == 'POST':
        expense.type = request.POST.get('expense-type')
        expense.description = request.POST.get('expense-description')
        expense.amount = request.POST.get('expense-amount')
        expense.save()
        # Create recent activity
        RecentActivity.objects.create(
            user=request.user,
            action=f"Updated expense: {expense.description} - ${expense.amount}",
            icon_class="fas fa-edit text-blue-500"
        )
        return redirect('admin_panel:dashboard')
    context = {'expense': expense}
    return render(request, 'admin_panel/update_expense.html', context)

@login_required
def delete_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    if request.method == 'POST':
        # Save to TrashBinEntry before deleting
        TrashBinEntry.objects.create(
            user=request.user,
            item_type='Expense',
            item_id=expense.id,
            item_data=json.loads(serialize('json', [expense]))[0]['fields'],
            deleted_at=timezone.now()
        )
        # Create recent activity before deleting
        RecentActivity.objects.create(
            user=request.user,
            action=f"Deleted expense: {expense.description} - ${expense.amount}",
            icon_class="fas fa-trash text-red-500"
        )
        expense.delete()
        return redirect('admin_panel:dashboard')
    context = {'expense': expense}
    return render(request, 'admin_panel/delete_expense.html', context)

@require_POST
@login_required
def add_transaction(request):
    date = request.POST.get('transaction-date')
    type_ = request.POST.get('transaction-type')
    description = request.POST.get('transaction-description')
    amount = request.POST.get('transaction-amount')
    if date and type_ and description and amount:
        transaction = Transaction.objects.create(
            date=date,
            type=type_,
            description=description,
            amount=amount,
        )
        # Create recent activity
        RecentActivity.objects.create(
            user=request.user,
            action=f"New transaction added: {description} - ${amount}",
            icon_class="fas fa-exchange-alt text-blue-500"
        )
    return redirect('admin_panel:dashboard')

def update_transaction(request, transaction_id):
    transaction = get_object_or_404(Transaction, id=transaction_id)
    if request.method == 'POST':
        transaction.date = request.POST.get('transaction-date')
        transaction.type = request.POST.get('transaction-type')
        transaction.description = request.POST.get('transaction-description')
        transaction.amount = request.POST.get('transaction-amount')
        transaction.save()
        return redirect('admin_panel:dashboard')
    context = {'transaction': transaction}
    return render(request, 'admin_panel/update_transaction.html', context)

def delete_transaction(request, transaction_id):
    transaction = get_object_or_404(Transaction, id=transaction_id)
    if request.method == 'POST':
        transaction.delete()
        return redirect('admin_panel:dashboard')
    context = {'transaction': transaction}
    return render(request, 'admin_panel/delete_transaction.html', context)

def fetch_tab_data(request, tab_name):
    if tab_name == 'trainees':
        data = Student.objects.filter(type='trainee').values()
    elif tab_name == 'internees':
        data = Student.objects.filter(type='internee').values()
    elif tab_name == 'iot':
        data = Student.objects.filter(type='iot').values()
    elif tab_name == 'sod':
        data = Student.objects.filter(type='sod').values()
    else:
        data = []

    return JsonResponse({'data': list(data)})

@login_required
def search(request):
    query = request.GET.get('q', '').strip()
    students = []
    employees = []
    expenses = []

    if query:
        students = Student.objects.filter(
            Q(name__icontains=query) |
            Q(type__icontains=query) |
            Q(category__icontains=query) |
            Q(program__icontains=query) |
            Q(level__icontains=query)
        )
        employees = Employee.objects.filter(
            Q(name__icontains=query) |
            Q(position__icontains=query) |
            Q(department__icontains=query)
        )
        expenses = Expense.objects.filter(
            Q(description__icontains=query) |
            Q(type__icontains=query)
        )

    context = {
        'query': query,
        'students': students,
        'employees': employees,
        'expenses': expenses,
    }
    return render(request, 'admin_panel/search_results.html', context)

@login_required
def export_employment_report_pdf(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 50
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, y, f"Employee Report: {employee.name}")
    y -= 30

    p.setFont("Helvetica", 12)
    p.drawString(50, y, f"Position: {employee.position}")
    y -= 20
    p.drawString(50, y, f"Department: {employee.department}")
    y -= 20
    p.drawString(50, y, f"Salary: {employee.salary}")
    y -= 30

    # Add more fields as needed...

    p.showPage()
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{employee.name}_report.pdf"'
    return response

@login_required
def export_employment_report_excel(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Report"

    ws.append(["Field", "Value"])
    ws.append(["Name", employee.name])
    ws.append(["Position", employee.position])
    ws.append(["Department", employee.department])
    ws.append(["Salary", str(employee.salary)])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{employee.name}_report.xlsx"'
    wb.save(response)
    return response

@login_required
def email_employment_report(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        recipient_email = request.POST.get('email')
        if recipient_email:
            subject = f"Employee Report: {employee.name}"
            message = render_to_string('admin_panel/email_employment_report.html', {'employee': employee})
            email = EmailMessage(subject, message, to=[recipient_email])
            email.content_subtype = "html"
            email.send()
            return JsonResponse({'success': True, 'message': 'Email sent successfully'})
        return JsonResponse({'success': False, 'message': 'Recipient email is required'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def financial_report_detail(request):
    total_salaries = Employee.objects.aggregate(Sum('salary'))['salary__sum'] or 0
    transport_expenses = Expense.objects.filter(type=Expense.TRANSPORT).aggregate(Sum('amount'))['amount__sum'] or 0
    other_expenses = Expense.objects.filter(type=Expense.OTHER).aggregate(Sum('amount'))['amount__sum'] or 0
    total_expenses = total_salaries + transport_expenses + other_expenses
    recent_transactions = Transaction.objects.order_by('-date')[:10]
    context = {
        'total_salaries': total_salaries,
        'transport_expenses': transport_expenses,
        'other_expenses': other_expenses,
        'total_expenses': total_expenses,
        'recent_transactions': recent_transactions,
    }
    return render(request, 'admin_panel/financial_report_detail.html', context)

@login_required
def export_financial_report_pdf(request):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 50
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, y, "Financial Report")
    y -= 30

    p.setFont("Helvetica", 12)
    total_salaries = Employee.objects.aggregate(Sum('salary'))['salary__sum'] or 0
    transport_expenses = Expense.objects.filter(type=Expense.TRANSPORT).aggregate(Sum('amount'))['amount__sum'] or 0
    other_expenses = Expense.objects.filter(type=Expense.OTHER).aggregate(Sum('amount'))['amount__sum'] or 0
    total_expenses = total_salaries + transport_expenses + other_expenses
    p.drawString(50, y, f"Total Salaries: ${total_salaries}")
    y -= 20
    p.drawString(50, y, f"Transport Expenses: ${transport_expenses}")
    y -= 20
    p.drawString(50, y, f"Other Expenses: ${other_expenses}")
    y -= 20
    p.drawString(50, y, f"Total Expenses: ${total_expenses}")
    y -= 30

    # Add more fields as needed...

    p.showPage()
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="financial_report.pdf"'
    return response

@login_required
def export_financial_report_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    ws.append(["Field", "Value"])
    total_salaries = Employee.objects.aggregate(Sum('salary'))['salary__sum'] or 0
    transport_expenses = Expense.objects.filter(type=Expense.TRANSPORT).aggregate(Sum('amount'))['amount__sum'] or 0
    other_expenses = Expense.objects.filter(type=Expense.OTHER).aggregate(Sum('amount'))['amount__sum'] or 0
    total_expenses = total_salaries + transport_expenses + other_expenses
    ws.append(["Total Salaries", str(total_salaries)])
    ws.append(["Transport Expenses", str(transport_expenses)])
    ws.append(["Other Expenses", str(other_expenses)])
    ws.append(["Total Expenses", str(total_expenses)])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="financial_report.xlsx"'
    wb.save(response)
    return response

@login_required
def email_financial_report(request):
    if request.method == 'POST':
        recipient_email = request.POST.get('email')
        if recipient_email:
            subject = "Financial Report"
            total_salaries = Employee.objects.aggregate(Sum('salary'))['salary__sum'] or 0
            transport_expenses = Expense.objects.filter(type=Expense.TRANSPORT).aggregate(Sum('amount'))['amount__sum'] or 0
            other_expenses = Expense.objects.filter(type=Expense.OTHER).aggregate(Sum('amount'))['amount__sum'] or 0
            total_expenses = total_salaries + transport_expenses + other_expenses
            context = {
                'total_salaries': total_salaries,
                'transport_expenses': transport_expenses,
                'other_expenses': other_expenses,
                'total_expenses': total_expenses,
            }
            message = render_to_string('admin_panel/email_financial_report.html', context)
            email = EmailMessage(subject, message, to=[recipient_email])
            email.content_subtype = "html"
            email.send()
            return JsonResponse({'success': True, 'message': 'Email sent successfully'})
        return JsonResponse({'success': False, 'message': 'Recipient email is required'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@csrf_exempt
@require_POST
@login_required
def bulk_delete_employees(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            employee_ids = data.get('employee_ids', [])
            if not employee_ids:
                return JsonResponse({'success': False, 'error': 'No employee IDs provided'})

            employees = Employee.objects.filter(id__in=employee_ids)
            deleted_count = 0
            for employee in employees:
                # Save to TrashBinEntry before deleting
                TrashBinEntry.objects.create(
                    user=request.user,
                    item_type='Employee',
                    item_id=employee.id,
                    item_data=json.loads(serialize('json', [employee]))[0]['fields'],
                    deleted_at=timezone.now()
                )
                deleted_count += 1

            # Bulk delete
            employees.delete()

            # Create recent activity
            RecentActivity.objects.create(
                user=request.user,
                action=f"Bulk deleted {deleted_count} employees",
                icon_class="fas fa-trash text-red-500"
            )

            return JsonResponse({'success': True, 'deleted_count': deleted_count})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required
def trash_bin(request):
    trash_entries = TrashBinEntry.objects.filter(user=request.user).order_by('-deleted_at')
    context = {
        'trash_entries': trash_entries,
    }
    return render(request, 'user_auth/trash_bin.html', context)
